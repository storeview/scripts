import xml.etree.ElementTree as ET
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border, Side, Font, PatternFill


"""
2022-1-19 添加新的功能，生成的 excel 文件，需要放置在 mm 同级目录下、

2022-2-25 将表格中的所有字段用 boarder 框起来

2023-1-30 程序优化点
    1. 第一行添加背景颜色, 加粗, 调整列宽
    2. pre, rem, intro等标记的加入, 进行处理
    3. 前置条件中, 如果几行内容是相同的则合并为同一行.

2023-2-01 
    将for循环转换为递归

2023-2-23
v1.0.5.0
    添加测试用例优先级字段.
"""

# 使用文件作为解析对象
print("#"*60)
freeplane_filename = input("请输入需要转换的 Freeplane 文件路径（或将文件拖入窗口中）:\n").replace('"','')
tree = ET.parse(freeplane_filename)

# 获得根节点
root = tree.getroot()
# Freeplane思维导图软件的根节点标签是 <map>
# root.tag

print("#"*60)
sheet_name = input("请输入测试用例生成的开始节点（该节点将作为工作蒲sheet的表名，下一级将作为【测试主模块】，若不输入任何值则以思维导图的根节点为起点）:\n")
# 选择测试用例生成的节点，该节点的下一级将作为【测试主模块】，该节点的名称将作为工作蒲 sheet 的表名
    #|-指定节点
    #|__test_main_module
    #|____test_module
    #|______test_content
    #|________test_step(or pre_condition or test_remark)
    #|__________expect_result


#print("------> " + root.tag)
# 获得该节点
top_node = root.find(".//node[@TEXT='"+sheet_name+"']")
# 如果获取不到值，说明 sheet_name 为空或者输入的节点名称不存在，此时应该将根节点作为默认值
if top_node == None:
    top_node = root.find("./node[@TEXT]")



# 新建工作蒲
wb = Workbook()
sheet = wb.active
sheet.title = top_node.get("TEXT")

# 设置表格样式（列宽、表头）
sheet.column_dimensions['A'].width=5
sheet.column_dimensions['B'].width=20
sheet.column_dimensions['C'].width=20
sheet.column_dimensions['D'].width=25
sheet.column_dimensions['E'].width=25
sheet.column_dimensions['F'].width=50
sheet.column_dimensions['G'].width=25
sheet.column_dimensions['H'].width=10
sheet.column_dimensions['I'].width=25
sheet.column_dimensions['J'].width=25


sheet.append(["编号", "测试主模块", "测试模块", "测试内容", "前置条件", "测试步骤", "预期结果", "测试结果", "备注", "异常情况说明", "优先级"])


# 解析mm文件, 并生成表格(代码可以进一步优化)

def deal_mm_file(node, level):
    if level == 4:
        level = level + 1
    text = node.get("TEXT")

    if text[0:2] == "tp":
        print("--------------------> 前置条件:\n " + text)
        one_row[4] = text[3:]
        return
    elif text[0:2] == "tr":
        print("--------------------> 备注:\n " + text)
        one_row[8] = text[3:]
        return
    elif text[0:2] == "te":
        print("--------------------> 异常情况说明:\n " + text)
        one_row[9] = text[3:]
        return
    elif len(text) == 2 and text[0:1] == "P":
        print("--------------------> 优先级:\n " + text)
        one_row[10] = text
        return 
    else:
        print("--------------------> level "+str(level)+" :\n " + text)
        one_row[level] = text

    # 递归下一层级
    for sub_node in node:
        deal_mm_file(sub_node, level+1)
    # 添加一行表格
    if level == 6:
        addNewNode2()
        # 清空最后三行的数据
        one_row[7] = ''
        one_row[8] = ''
        one_row[9] = ''
        one_row[10] = ''

    # 清空本次递归填写的数据
    if level != 7:
        one_row[level] = ''

def addNewNode2():
    print("--------------------> 正在输出一行表格 <--------------------")
    sheet.append([one_row[0].strip(), one_row[1].strip(), one_row[2].strip(), one_row[3].strip(), one_row[4].strip(), "\n"+one_row[5].strip()+"\n", one_row[6].strip(), one_row[7].strip(), one_row[8].strip(), one_row[9].strip(), one_row[10].strip()])
                    



# 将上面复杂的代码, 转换为一个递归进行处理
# 递归中包含这样的逻辑: 
#   当此节点的子节点长度为0时,代表此节点是最后一个了, 此时结束循环.

one_row = ['', '', '', '', '', '', '', '', '', '', '']
# 测试主模块
for test_main_module in top_node:
    if test_main_module.tag != "node":
        continue
    deal_mm_file(test_main_module, 1)
    # 清空
    one_row = ['', '', '', '', '', '', '', '', '', '', '']



# 表格生成完成，进行格式处理
thin = Side(border_style="thin", color="000000")

i = 1
# 对所有单元格进行处理(横排竖排对齐, 以及边框)
while sheet.cell(i, 1).value != None:
    for j in range(1, 12):
        if i== 1:
            # 上下左右居中对齐，自动换行
            sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        elif j == 1 or j == 2 or j == 3 or j == 4:
            sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        else:
            sheet.cell(i, j).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        # 边框
        sheet.cell(i, j).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    i += 1

# 设置背景颜色和加粗
k = 1
ft = Font(bold=True)
while sheet.cell(1, k).value != None:
    sheet.cell(1, k).font = ft
    sheet.cell(1, k).fill = PatternFill("solid", fgColor="5b9bd5")
    k+=1

# 对2-5列的表格进行相邻行内容相同归并处理
for col in range (2, 6):
    row = 2
    start_row = 2
    end_row = start_row
    start_value = sheet.cell(row, col).value 
    #if start_value == None or start_value == "":
    #    continue
    while sheet.cell(row+1, col).value != None:
        print(str(row+1) +", " + str(col))
        if sheet.cell(row + 1, col).value == start_value and start_value != "":
            end_row = row + 1
        else:
            sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            start_row = row + 1
            end_row = start_row
            start_value = sheet.cell(start_row, col).value
        row += 1
    sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)


# 需要使用正则表达式，匹配文件名，并复制给新文件
m = re.search(r'^(.*)\\([^\\]+).mm$',freeplane_filename)
file_parent_dir = m.group(1)
filename = m.group(2)

wb.save(file_parent_dir + "\\"+str(int(round(time.time() * 1000)))+filename+".xlsx")   

print("#"*60)
print("OK")
