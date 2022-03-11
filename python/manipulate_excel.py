'''
渲染 Buglist 表格文件

- 目前存在 bug：如果 excel 中存储了其他对象（例如图片），会顺带着把图片也清除了


'''



import openpyxl
from openpyxl.reader import excel
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 获取 Excel 文件
path = input("请将需要修改样式的 buglist 文件拖入命令行窗口，再敲击回车\n")
excelFile = openpyxl.load_workbook(path)

redFont = Font(name="等线",size=11,bold=True,i=False,color="FF0000")
orangeFont = Font(name="等线",size=11,bold=True,i=False,color="C65911")
greenFont = Font(name="等线",size=11,bold=True,i=False,color="00B05D")
Color = ['FF0000', '00B05D', 'BFBFBF']
redFill = PatternFill('solid', fgColor=Color[0])
greenFill = PatternFill('solid', fgColor=Color[1])
grayFill = PatternFill('solid', fgColor=Color[2])
# 1. 遍历所有的 sheet
for sheet in excelFile.worksheets:
    # 2. 遍历所有的行
    for row in sheet.iter_rows():
        # 3. 将所有的等于 pass、block、fail 文字的表格（cell）修改样式
        for cell in row:
            if cell.value == "pass":
                cell.font = greenFont 
            elif cell.value == "block":
                cell.font = orangeFont
            elif cell.value == '已修复':
                cell.fill = greenFill
            elif cell.value == '新增BUG':
                cell.fill = redFill
            elif cell.value == '未修复':
                cell.fill = grayFill
            elif cell.value == "fail" or cell.value == '严重':
                cell.font = redFont

excelFile.save(path)
input("OK! 按 Enter 键结束")



# 创建一个全新的 Excel 文件
newExcelFile = openpyxl.Workbook()
newExcelFile.create_chartsheet(index=0, title="香蕉")
newExcelFile.create_chartsheet(index=1, title="苹果")
# newExcelFile.save("123.xlsx")
