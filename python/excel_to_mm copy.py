"""
20230628 解析『平台用例』导出的Excel文件
"""


from openpyxl import load_workbook
import re

excel_filename = "D:\\Windows数据移动到此文件夹\\Desktop\\测试用例导出.xlsx"
wb = load_workbook(excel_filename)
sheet_name = "Sheet1"
sheet = wb[sheet_name]

datas = list(sheet.rows)
title_zh = [i.value for i in datas[0]]
title_en = [i.value for i in datas[1]]

class OpmlTree():
    def __init__(self, root_name):
        self.str = "<map>\n<node TEXT='"+root_name+"' COLOR='#000000' FOLDED='false'>\n"
        self.depth = 0
        self.end_node_num = 0
    def build_category_level_node(self, val):
        self.str += "<node TEXT='"+val+"' COLOR='#000000' FOLDED='true'>\n"
        self.end_node_num += 1
    def build_middle_level_node(self, val):
        if val != None:
            self.str += "<node TEXT='"+val+"' COLOR='#000000' FOLDED='true'/>\n"
        else:
            self.str += "<node TEXT='' COLOR='#000000' FOLDED='true'/>\n"
    def build_end_node(self, target_num):
        while self.end_node_num >= target_num:
            self.str += "</node>\n"
            self.end_node_num -= 1
    def output_str(self):
        self.build_end_node(0)
        self.str += "</map>"
        return self.str
    def get_depth(self):
        return self.depth
    def set_depth(self, depth):
        self.depth = depth

opml_file = OpmlTree("Root")

#[{a:a1, b:b1, c:c1}, {d:d1, e:e1, f:f1}]
for i in datas[2:]:
    values = [c.value for c in i]
    row = dict(zip(title_en, values))
    if row['isFeature'] != None and row['isFeature'].strip() == "true":
        #print(str(len(row['Depth'].encode()))+":"+row["Depth"])
        if len(row['Depth'].encode()) <= opml_file.get_depth():
            opml_file.build_end_node(len(row['Depth'].encode()))
        opml_file.build_category_level_node(row['Feature_Name'])
    else:
        opml_file.build_category_level_node(row['Testcase_Number'])
        opml_file.build_category_level_node(row['Testcase_Name'])
        opml_file.build_category_level_node(row['Testcase_Prepare'])
        opml_file.build_category_level_node(row['TestCase_TestStep'])
        opml_file.build_category_level_node(row['Testcase_Expect'])
        opml_file.build_end_node(len(row['Depth'].encode()))
    opml_file.set_depth(len(row['Depth'].encode()))



m = re.search(r'^(.*)\\([^\\]+).xlsx$',excel_filename)
file_parent_dir = m.group(1)
filename = m.group(2)

f = open(file_parent_dir + "\\temp_"+filename+".mm", 'w', encoding='utf-8')
f.write(opml_file.output_str())
f.close()

'''
root_name = input("请输入思维导图根节点名称（如果不输入，则以 sheet 表格名称作为根节点）：\n")


m_list = sheet.merged_cells  #合并单元格的位置信息，可迭代对象
cr = []
for m_area in m_list:
    # 合并单元格的起始行坐标、终止行坐标。。。。，
    r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
    # 纵向合并单元格的位置信息提取出
    if r2 - r1 > 0:
        cr.append((r1, r2, c1, c2))

def copyData(start_row, end_row, start_col, end_col):
    """将左上角的第一个数据，复制到所有合并的单元格中"""
    value = sheet.cell(start_row, start_col).value
    for i in range(start_row, end_row+1):
        for j in range(start_col, end_col+1):
            sheet.cell(i, j ).value = value

# 纵向单元格取消合并
for r in cr:
    sheet.unmerge_cells(start_row=r[0], end_row=r[1],
                            start_column=r[2], end_column=r[3])
    # 填充数据
    copyData(r[0], r[1], r[2], r[3])




class Node():
    """
        N叉树的基本节点
    """
    # 初始化一个节点
    def __init__(self, val = None):
        self.val = val
        self.children = []
    # 添加一个子孩子
    def add_child(self, node):
        self.children.append(node) 
        

class TestCateTree():
    """
        存放测试用例的多叉树
    """
    # 初始化一颗N叉树
    def __init__(self, val):
        self.root = Node(val)
        self.len = 0
        
    # 从指定节点开始查找【含有特定值】的节点，如果找到则返回该节点，没有找到则创建该节点
    def search(self, node, target_val):
        # 查找是否存在该子节点，存在则返回该节点
        for index, child in enumerate(node.children):
            if child.val == target_val:
                return node.children[index]
        # 不存在则添加该节点，并返回新添加的节点
        node.add_child(Node(target_val))
        self.len += 1
        return node.children[-1]

    def addNode(self, node, new_node_val):
        for index, child in enumerate(node.children):
            if child.val == new_node_val:
                return
        node.add_child(Node(new_node_val))
    

if root_name == None or root_name == "":
    root_name = sheet_name
tree = TestCateTree(root_name)


# 以行为单位，遍历 sheet 表格的所有行，构建 Python N 叉树
i = 2
while sheet.cell(i, 2).value != None:
    # 获取每一行各列单元格的数值，需要考虑到遍历出来的 Excel 表格可能存在空值
    test_main_model = sheet.cell(i, 2).value
    test_model = sheet.cell(i, 3).value
    test_content = sheet.cell(i, 4).value
    pre_condition = sheet.cell(i, 5).value
    test_method = sheet.cell(i, 6).value
    expect_result = sheet.cell(i, 7).value
    test_remark = sheet.cell(i, 8).value

    parent_node = tree.root
    parent_node = tree.search(parent_node, test_main_model)
    parent_node = tree.search(parent_node, test_model)
    if test_content != None and test_content != "":
        parent_node = tree.search(parent_node, "tc: "+test_content)
        if pre_condition != None and pre_condition != "":
            tree.addNode(parent_node, "tp: "+pre_condition)
        if test_remark != None and test_remark != "":
            tree.addNode(parent_node, "tr: "+str(test_remark))
        if test_method != None and test_method != "":
            parent_node = tree.search(parent_node, "ts: "+test_method)
            parent_node = tree.search(parent_node, expect_result)
    i += 1

    

# ↑对 Excel 文件的解析已经完成，并且声称了 Python N 叉树
#       下面是对 Python N 叉树的先序遍历，并生成对应的 mm 文件
def pre_order(node, level):
    if node.val == None:
        return ""
    tmp_str = ""

    # 回车符 &#xa;
    val = node.val.replace("\n", "&#xa;")
    
    if level == 1:
        tmp_str = "<node TEXT='"+val+"' FOLDED='false'>\n"
    elif level == 2:
        tmp_str = "<node TEXT='"+val+"' FOLDED='true'>\n"
    elif level == 3:
        tmp_str = "<node TEXT='"+val+"' FOLDED='true'>\n"
    elif level == 4:
        tmp_str = "<node TEXT='"+val+"' >\n"
    elif level == 5:
        tmp_str = "<node TEXT='"+val+"' >\n"
    elif level == 6:
        tmp_str = "<node TEXT='"+val+"'/>\n"
    else:
        tmp_str = "<node TEXT='"+val+"'>\n"

    for index, sub_node in enumerate(node.children):
        tmp_str += pre_order(sub_node, level+1)
    
    if level != 6:
        tmp_str += "</node>\n"
    
    return tmp_str


opml_str = "<map >\n"
opml_str += pre_order(tree.root, 1)
opml_str += "</map>"


m = re.search(r'^(.*)\\([^\\]+).xlsx$',excel_filename)
file_parent_dir = m.group(1)
filename = m.group(2)

f = open(file_parent_dir + "\\temp_"+filename+".mm", 'w', encoding='utf-8')
f.write(opml_str)
f.close()

print("OK")

'''